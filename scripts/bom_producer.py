#!/usr/bin/env python3
# scripts/bom_producer.py
# Transform technical BOM (JSONL) → production BOM (CSV/XLSX)

import sys, json, csv, argparse, math
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("WARNING: openpyxl not installed; XLSX export disabled", file=sys.stderr)

p = argparse.ArgumentParser(description="Transform technical BOM (JSONL) → production BOM (CSV/XLSX)")
p.add_argument("--jsonl", required=True, help="Input JSONL file (from render_bom.py)")
p.add_argument("--parts", required=True, help="Parts catalog CSV (data/parts.csv)")
p.add_argument("--csv", default="", help="Output CSV file")
p.add_argument("--xlsx", default="", help="Output XLSX file (requires openpyxl)")
p.add_argument("--debug", action="store_true", help="Print debug info")
args = p.parse_args()

# --- Load parts catalog ---
parts_catalog = defaultdict(lambda: defaultdict(dict))  # [category][enum_value] = {part_no, material_code, ...}
with open(args.parts, encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        category = row["category"]
        enum_val = row["enum_value"]
        parts_catalog[category][enum_val] = {
            "material_code": row.get("material_code", ""),
            "part_no": row.get("part_no", ""),
            "description": row.get("description", ""),
            "unit": row.get("unit", ""),
            "supplier": row.get("supplier", ""),
        }

if args.debug:
    sys.stderr.write(f"Loaded {len(parts_catalog)} categories\n")

# --- Load technical BOM ---
bom_records = []
with open(args.jsonl, encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if line:
            try:
                rec = json.loads(line)
                bom_records.append(rec)
            except json.JSONDecodeError as e:
                sys.stderr.write(f"JSON error: {e}\n")
                continue

if args.debug:
    sys.stderr.write(f"Loaded {len(bom_records)} BOM records\n")

# --- Transform to production BOM ---
def lookup_part(category, enum_val):
    """Lookup part info from catalog; return part_no or enum_val if not found."""
    if category in parts_catalog and enum_val in parts_catalog[category]:
        return parts_catalog[category][enum_val]
    return {"part_no": f"UNMAPPED-{enum_val}", "material_code": enum_val, "description": "UNMAPPED", "unit": "?", "supplier": "?"}

def calculate_surface_area(D, L):
    """Calculate surface area of cylinder: π*D*L (outer surface)."""
    return round(math.pi * D * L / 1_000_000, 4)  # mm² → m²

def calculate_cut_length(D, L, rings_count):
    """Rough estimate: perimeter * L + ring cuts."""
    perim = math.pi * D
    ring_cuts = rings_count * perim
    total = (perim * L + ring_cuts) / 1_000  # mm → m
    return round(total, 2)

production_bom = []
for tech_bom in bom_records:
    rec = {
        "product": tech_bom.get("product", ""),
        "version": tech_bom.get("version", ""),
        "bom_tag": tech_bom.get("bom_tag", ""),
    }
    
    # Basic dimensions
    L = tech_bom.get("L", 0)
    D = tech_bom.get("D", 0)
    t = tech_bom.get("t", 0)
    rec["length_mm"] = L
    rec["diameter_mm"] = D
    rec["thickness_mm"] = t
    
    # Material & codes
    medium = tech_bom.get("medium", "")
    material_info = lookup_part("material", medium)
    rec["material"] = medium
    rec["material_code"] = material_info.get("material_code", "")
    rec["material_part_no"] = material_info.get("part_no", "")
    rec["material_supplier"] = material_info.get("supplier", "")
    
    # Top
    top = tech_bom.get("top", "")
    open_top = tech_bom.get("open_top", False)
    if not open_top and top:
        top_info = lookup_part("top", top)
        rec["top_type"] = top
        rec["top_part_no"] = top_info.get("part_no", "")
        rec["top_supplier"] = top_info.get("supplier", "")
    else:
        rec["top_type"] = "open" if open_top else ""
        rec["top_part_no"] = ""
        rec["top_supplier"] = ""
    
    # Bottom
    bottom = tech_bom.get("bottom", "")
    bottom_opt = tech_bom.get("bottom_opt", "")
    bottom_info = lookup_part("bottom", bottom)
    rec["bottom_type"] = bottom
    rec["bottom_part_no"] = bottom_info.get("part_no", "")
    
    if bottom_opt and bottom_opt != "zonder":
        bottom_opt_info = lookup_part("bottom_opt", bottom_opt)
        rec["bottom_option"] = bottom_opt
        rec["bottom_option_part_no"] = bottom_opt_info.get("part_no", "")
    else:
        rec["bottom_option"] = ""
        rec["bottom_option_part_no"] = ""
    
    # Rings
    rings = tech_bom.get("rings", [])
    ring_count = len(rings) if rings else 0
    ring_w = tech_bom.get("ring_w", 0)
    ring_t = tech_bom.get("ring_t", 0)
    rec["ring_count"] = ring_count
    rec["ring_width_mm"] = ring_w
    rec["ring_thickness_mm"] = ring_t
    
    # Reinforcement
    reinforce_enable = tech_bom.get("reinforce", False)
    reinforce_side = tech_bom.get("rein_side", "")
    reinforce_spans = tech_bom.get("rein_spans", [])
    rec["reinforce_enabled"] = "Yes" if reinforce_enable else "No"
    
    if reinforce_enable and reinforce_side:
        reinf_info = lookup_part("reinforcement", reinforce_side)
        rec["reinforcement_type"] = reinforce_side
        rec["reinforcement_part_no"] = reinf_info.get("part_no", "")
        # Calculate total reinforcement length (sum of spans)
        total_reinf_length = sum(span[1] - span[0] for span in reinforce_spans if len(span) == 2)
        rec["reinforcement_length_mm"] = total_reinf_length
    else:
        rec["reinforcement_type"] = ""
        rec["reinforcement_part_no"] = ""
        rec["reinforcement_length_mm"] = 0
    
    # Productzijde
    productzijde = tech_bom.get("productzijde", "")
    rec["productzijde"] = productzijde
    
    # Calculated fields
    rec["surface_area_m2"] = calculate_surface_area(D, L)
    rec["cut_length_estimate_m"] = calculate_cut_length(D, L, ring_count)
    
    production_bom.append(rec)

if args.debug:
    sys.stderr.write(f"Produced {len(production_bom)} production records\n")

# --- Export CSV ---
if args.csv:
    out_csv = Path(args.csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    
    # Define column order
    fieldnames = [
        "product", "version", "bom_tag",
        "material", "material_code", "material_part_no", "material_supplier",
        "length_mm", "diameter_mm", "thickness_mm",
        "top_type", "top_part_no", "top_supplier",
        "bottom_type", "bottom_option", "bottom_part_no", "bottom_option_part_no",
        "ring_count", "ring_width_mm", "ring_thickness_mm",
        "reinforce_enabled", "reinforcement_type", "reinforcement_part_no", "reinforcement_length_mm",
        "productzijde",
        "surface_area_m2", "cut_length_estimate_m",
    ]
    
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, restval="")
        writer.writeheader()
        for rec in production_bom:
            writer.writerow(rec)
    
    print(f"✓ CSV exported to {out_csv}")

# --- Export XLSX ---
if args.xlsx:
    if not XLSX_AVAILABLE:
        sys.stderr.write("ERROR: openpyxl required for XLSX export. Install: pip install openpyxl\n")
        sys.exit(1)
    
    out_xlsx = Path(args.xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    
    # Define columns
    fieldnames = [
        "product", "version", "bom_tag",
        "material", "material_code", "material_part_no",
        "length_mm", "diameter_mm", "thickness_mm",
        "top_type", "top_part_no",
        "bottom_type", "bottom_option",
        "ring_count", "ring_width_mm", "ring_thickness_mm",
        "reinforce_enabled", "reinforcement_type", "reinforcement_length_mm",
        "productzijde",
        "surface_area_m2", "cut_length_estimate_m",
    ]
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_idx, fieldname in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=fieldname)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, rec in enumerate(production_bom, 2):
        for col_idx, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(fieldname, ""))
            # Number formatting
            if "mm" in fieldname or "m2" in fieldname:
                cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-size columns
    for col_idx, fieldname in enumerate(fieldnames, 1):
        max_length = len(fieldname) + 2
        for row in ws.iter_rows(min_row=2, max_row=len(production_bom) + 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length, 30)
    
    wb.save(out_xlsx)
    print(f"✓ XLSX exported to {out_xlsx}")

print("✓ BOM production complete")
